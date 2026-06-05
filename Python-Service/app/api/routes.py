import json
import os
import tempfile
import shutil
import urllib.parse 
import pandas as pd
from openpyxl import load_workbook
from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Request, Query, Header 
from fastapi.responses import RedirectResponse
import httpx 
from pathlib import Path
from dotenv import load_dotenv

from app.services.validator_service import process_validation_batch
from app.services.crm_metadata_service import CrmMetadataService

# Load environment variables from .env file
current_file = Path(__file__).resolve()
possible_paths = [
    current_file.parent.parent.parent / ".env",  # Python-Service/ folder
    Path.cwd() / ".env",                          # Active Terminal Workspace folder
    Path.cwd() / "Python-Service" / ".env"        # Sub-folder targeting
]

# Loads the first valid environment configuration configuration file it encounters
for path in possible_paths:
    if path.exists():
        load_dotenv(dotenv_path=path)
        break

router = APIRouter()

# =========================================================
# OAUTH CONFIGURATION (Loaded safely from .env)
# =========================================================
ANGULAR_FRONTEND_URL = os.getenv("ANGULAR_FRONTEND_URL", "http://localhost:4200/connection")
FASTAPI_BACKEND_URL = os.getenv("FASTAPI_BACKEND_URL", "http://localhost:8000")

SF_CLIENT_ID = os.getenv("SF_CLIENT_ID", "").strip()
SF_CLIENT_SECRET = os.getenv("SF_CLIENT_SECRET", "").strip()
SF_REDIRECT_URI = f"{FASTAPI_BACKEND_URL}/api/auth/salesforce/callback"

ZD_CLIENT_ID = os.getenv("ZD_CLIENT_ID", "").strip()
ZD_CLIENT_SECRET = os.getenv("ZD_CLIENT_SECRET", "").strip()
ZD_REDIRECT_URI = f"{FASTAPI_BACKEND_URL}/api/auth/zendesk/callback"
#Zoho ClientId and secret and URI
ZOHO_CLIENT_ID = os.getenv("ZOHO_CLIENT_ID", "").strip()
ZOHO_CLIENT_SECRET = os.getenv("ZOHO_CLIENT_SECRET", "").strip()
ZOHO_REGIONS = {
    "us": "https://accounts.zoho.com",
    "in": "https://accounts.zoho.in",
    "eu": "https://accounts.zoho.eu",
    "au": "https://accounts.zoho.com.au",
    "jp": "https://accounts.zoho.jp",
    "ca": "https://accounts.zohocloud.ca",
    "sa": "https://accounts.zoho.sa",
    "uk": "https://accounts.zoho.uk",
    "cn": "https://accounts.zoho.com.cn"
}
FASTAPI_BACKEND_URL = os.getenv("FASTAPI_BACKEND_URL", "http://localhost:8000").rstrip("/")
ANGULAR_FRONTEND_URL = os.getenv("ANGULAR_FRONTEND_URL", "http://localhost:4200")
ZOHO_REDIRECT_URI = f"{FASTAPI_BACKEND_URL}/api/auth/zoho/callback"


# =========================================================
# ROUTE: DYNAMIC OAUTH LOGIN INITIALIZATION
# =========================================================
@router.get("/api/auth/{crm_id}/login")
async def crm_oauth_login(
    crm_id: str, 
    side: str = Query(...),  # 'source' or 'target'
    subdomain: str = Query(None), # Required only for zendesk
    region: str = Query(None) # Capture region from Angular UI (e.g., 'US', 'IN', 'EU') for zoho
):
    crm_id_lower = crm_id.lower()
    
    if crm_id_lower == "salesforce":
        params = {
            "client_id": SF_CLIENT_ID,
            "redirect_uri": SF_REDIRECT_URI,
            "response_type": "code",
            "state": side
        }
        query_string = urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
        auth_url = f"https://login.salesforce.com/services/oauth2/authorize?{query_string}"
        return RedirectResponse(auth_url)

    elif crm_id_lower == "zendesk":
        if not subdomain:
            raise HTTPException(status_code=400, detail="Zendesk requires a subdomain parameter.")
        
        params = {
            "client_id": ZD_CLIENT_ID,
            "redirect_uri": ZD_REDIRECT_URI,
            "response_type": "code",
            "state": f"{side}:{subdomain}",
            "scope": "read write",
        }
        query_string = urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
        auth_url = f"https://{subdomain.strip()}.zendesk.com/oauth/authorizations/new?{query_string}"
        return RedirectResponse(auth_url)
    elif crm_id_lower == "zoho":
        if not ZOHO_CLIENT_ID:
            raise HTTPException(status_code=500, detail="ZOHO_CLIENT_ID is missing from environment variables.")
            
        reg_key = (region or "US").lower().strip()
        base_accounts_url = ZOHO_REGIONS.get(reg_key, ZOHO_REGIONS["us"])
        
        scopes = [
            "ZohoCRM.modules.ALL",
            "ZohoCRM.bulk.READ",
            "ZohoCRM.settings.FIELDS.READ",
            "ZohoCRM.settings.modules.READ",
            "ZohoCRM.coql.READ"
        ]
        
        # CORRECTED FIXED PARAMS BLOCK BELOW:
        params = {
            "scope": ",".join(scopes),
            "client_id": ZOHO_CLIENT_ID,
            "response_type": "code",
            "access_type": "offline",
            "redirect_uri": ZOHO_REDIRECT_URI,
            "prompt": "consent",
            "state": f"{side}:{reg_key}"
        }
        
        query_string = urllib.parse.urlencode(params)
        auth_url = f"{base_accounts_url}/oauth/v2/auth?{query_string}"
        return RedirectResponse(auth_url)
    raise HTTPException(status_code=400, detail=f"CRM engine '{crm_id}' is not yet supported via OAuth.")


# =========================================================
# ROUTE: SALESFORCE OAUTH CALLBACK
# =========================================================
@router.get("/api/auth/salesforce/callback")
async def salesforce_callback(code: str, state: str):
    # 👇 CRITICAL FIX: Added verify=False here
    async with httpx.AsyncClient(verify=False) as client:
        response = await client.post("https://login.salesforce.com/services/oauth2/token", data={
            "grant_type": "authorization_code",
            "code": code,
            "client_id": SF_CLIENT_ID,
            "client_secret": SF_CLIENT_SECRET,
            "redirect_uri": SF_REDIRECT_URI
        })
        
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Failed to retrieve token from Salesforce.")
            
        token_data = response.json()
        access_token = token_data.get("access_token")
        instance_url = token_data.get("instance_url", "")
        
        safe_instance_url = urllib.parse.quote(instance_url)

    return RedirectResponse(f"{ANGULAR_FRONTEND_URL}?connected_side={state}&crm=salesforce&access_token={access_token}&instance_url={safe_instance_url}")


# =========================================================
# ROUTE: ZENDESK OAUTH CALLBACK
# =========================================================
@router.get("/api/auth/zendesk/callback")
async def zendesk_callback(code: str, state: str):
    try:
        side, subdomain = state.split(":")
    except ValueError:
        raise HTTPException(status_code=400, detail="State parameter verification corruption.")

    # 👇 CRITICAL FIX: Added verify=False here
    async with httpx.AsyncClient(verify=False) as client:
        response = await client.post(f"https://{subdomain}.zendesk.com/oauth/tokens", json={
            "grant_type": "authorization_code",
            "code": code,
            "client_id": ZD_CLIENT_ID,
            "client_secret": ZD_CLIENT_SECRET,
            "redirect_uri": ZD_REDIRECT_URI,
            "scope": "read write"
        })
        
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Failed to retrieve token from Zendesk.")
            
        token_data = response.json()

    access_token = token_data.get("access_token")
    return RedirectResponse(f"{ANGULAR_FRONTEND_URL}?connected_side={side}&crm=zendesk&access_token={access_token}")

# =========================================================
# ROUTE: ZOHO OAUTH CALLBACK
# =========================================================
@router.get("/api/auth/zoho/callback")
async def zoho_callback(code: str, state: str, request: Request):
    try:
        side, reg_key = state.split(":") 
    except ValueError:
        side = state
        reg_key = "us"
    base_accounts_url = ZOHO_REGIONS.get(reg_key.lower(), ZOHO_REGIONS["us"])
    accounts_server = request.query_params.get("accounts-server", base_accounts_url)
    
    # 👇 CRITICAL FIX: Added verify=False here to match your other routes!
    async with httpx.AsyncClient(verify=False) as client:
        response = await client.post(f"{accounts_server}/oauth/v2/token", data={
            "grant_type": "authorization_code",
            "code": code,
            "client_id": ZOHO_CLIENT_ID,
            "client_secret": ZOHO_CLIENT_SECRET,
            "redirect_uri": ZOHO_REDIRECT_URI
        })
        
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Failed to retrieve token from Zoho CRM.")
            
        token_data = response.json()
        access_token = token_data.get("access_token")
        api_domain = token_data.get("api_domain", "https://www.zohoapis.com")
        
        safe_api_domain = urllib.parse.quote(api_domain)
        safe_accounts_server = urllib.parse.quote(accounts_server)

    return RedirectResponse(
        f"{ANGULAR_FRONTEND_URL}?connected_side={side}&crm=zoho&access_token={access_token}&api_domain={safe_api_domain}&accounts_server={safe_accounts_server}"
    )

# ==========================================
# ROUTE 1: FAST HEADER EXTRACTION
# ==========================================
@router.post("/api/python/extract-headers")
async def extract_headers(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename)[1].lower()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    temp_file_name = temp_file.name
    temp_file.close() 
    
    try:
        with open(temp_file_name, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        sheets = []
        headers_map = {}

        if ext == '.csv':
            df = pd.read_csv(temp_file_name, nrows=0)
            sheets = ["Sheet1"]
            headers_map["Sheet1"] = df.columns.tolist()
            
        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(temp_file_name, read_only=True, data_only=True)
            sheets = wb.sheetnames
            for sheet in sheets:
                ws = wb[sheet]
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
                headers_map[sheet] = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(first_row)]
            wb.close()
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format.")

        return {
            "sheets": sheets,
            "headersMap": headers_map
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        file.file.close()
        if os.path.exists(temp_file_name):
            try:
                os.remove(temp_file_name)
            except PermissionError:
                pass


# ==========================================
# ROUTE 2: MASSIVE DATA VALIDATION (CHUNKS)
# ==========================================
@router.post("/api/python/validate")
async def validate_batch(
    file: UploadFile = File(...),
    config: str = Form(...) 
):
    try:
        payload = json.loads(config)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid configuration format")

    mappings = payload.get("mappings", [])
    dedupe_key = payload.get("dedupeKey", "")
    sheet_name = payload.get("sheetName", "")
    sf_rules = payload.get("sfRules", {})
    date_format = payload.get("dateFormat", "")

    ext = os.path.splitext(file.filename)[1].lower()

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    temp_file_name = temp_file.name
    temp_file.close()

    try:
        with open(temp_file_name, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        total_count = 0
        total_valid = 0
        total_invalid_count = 0
        total_duplicates = 0
        all_invalid_records = []
        all_valid_records = []

        if ext == '.csv':
            chunk_iterator = pd.read_csv(temp_file_name, chunksize=10000)
            
            for chunk_df in chunk_iterator:
                chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                chunk_records = chunk_df.to_dict(orient="records")
                
                result = process_validation_batch(
                    records=chunk_records, mappings=mappings, dedupe_key=dedupe_key, 
                    sf_rules=sf_rules, date_format=date_format
                )
                
                total_count += result["stats"]["total"]
                total_valid += result["stats"]["valid"]
                total_invalid_count += result["stats"]["invalid"]
                total_duplicates += result["stats"]["duplicates"]
                all_invalid_records.extend(result["invalidRecords"])
                all_valid_records.extend(result["validRecords"])

        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(temp_file_name, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            headers_raw = next(rows_iter, [])
            headers = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(headers_raw)]

            chunk_records = []
            
            for row in rows_iter:
                if not any(row): continue 
                
                chunk_records.append(dict(zip(headers, row)))
                
                if len(chunk_records) == 10000:
                    chunk_df = pd.DataFrame(chunk_records)
                    chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                    
                    result = process_validation_batch(
                        records=chunk_df.to_dict(orient="records"), mappings=mappings, dedupe_key=dedupe_key,  
                        sf_rules=sf_rules, date_format=date_format
                    )
                    
                    total_count += result["stats"]["total"]
                    total_valid += result["stats"]["valid"]
                    total_invalid_count += result["stats"]["invalid"]
                    total_duplicates += result["stats"]["duplicates"]
                    all_invalid_records.extend(result["invalidRecords"])
                    all_valid_records.extend(result["validRecords"])
                    
                    chunk_records = []
            
            if chunk_records:
                chunk_df = pd.DataFrame(chunk_records)
                chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                
                result = process_validation_batch(
                    records=chunk_df.to_dict(orient="records"), mappings=mappings, dedupe_key=dedupe_key,  
                    sf_rules=sf_rules, date_format=date_format
                )
                
                total_count += result["stats"]["total"]
                total_valid += result["stats"]["valid"]
                total_invalid_count += result["stats"]["invalid"]
                total_duplicates += result["stats"]["duplicates"]
                all_invalid_records.extend(result["invalidRecords"])
                all_valid_records.extend(result["validRecords"])
                
            wb.close()

        return {
            "stats": {
                "total": total_count,
                "valid": total_valid,
                "invalid": total_invalid_count,
                "duplicates": total_duplicates
            },
            "invalidRecords": all_invalid_records,
            "validRecords": all_valid_records
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        file.file.close()
        if os.path.exists(temp_file_name):
            try:
                os.remove(temp_file_name)
            except PermissionError:
                pass


# ==========================================
# ROUTE 3: QUICK RE-VALIDATION (JSON)
# ==========================================
@router.post("/api/python/revalidate")
async def revalidate_batch_json(request: Request):
    payload = await request.json()
    records = payload.get("records", [])
    mappings = payload.get("mappings", [])
    dedupe_key = payload.get("dedupeKey", "")
    sf_rules = payload.get("sfRules", {})
    date_format = payload.get("dateFormat", "")

    result = process_validation_batch(
        records=records, 
        mappings=mappings, 
        dedupe_key=dedupe_key,  
        sf_rules=sf_rules,
        date_format=date_format
    )
    
    return result


# =========================================================
# ROUTE: FETCH DYNAMIC CRM CORE ENTITIES / OBJECT LIST
# =========================================================
@router.get("/api/metadata/{crm_id}/objects")
async def get_crm_objects(
    crm_id: str,
    sf_token: str = Header(None, alias="sf-token"),
    sf_instance_url: str = Header(None, alias="sf-instance-url"),
    zd_token: str = Header(None, alias="zd-token"),
    zd_subdomain: str = Header(None, alias="zd-subdomain"),
    zoho_token: str = Header(None, alias="zoho-token"),
    zoho_api_domain: str = Header(None, alias="zoho-api-domain")
):
    crm_lower = crm_id.lower()
    
    if crm_lower == "salesforce":
        return await CrmMetadataService.fetch_salesforce_objects(sf_token, sf_instance_url)
        
    elif crm_lower == "zendesk":
        return await CrmMetadataService.fetch_zendesk_objects(zd_token, zd_subdomain)
        
    elif crm_lower == "zoho":
        # 1. Sanitize the Zoho Domain Protocol
        if zoho_api_domain and not zoho_api_domain.startswith(("http://", "https://")):
            zoho_api_domain = f"https://{zoho_api_domain}"
        base_url = zoho_api_domain.rstrip('/') if zoho_api_domain else "https://www.zohoapis.com"

        # 2. Execute with verify=False for SSL safety
        async with httpx.AsyncClient(verify=False, timeout=30.0) as client:
            res = await client.get(f"{base_url}/crm/v6/settings/modules", headers={
                "Authorization": f"Zoho-oauthtoken {zoho_token}"
            })
            if res.status_code != 200: 
                    raise HTTPException(status_code=res.status_code, detail=f"Zoho API Error: {res.text}")
            data = res.json()
            zoho_objects = []
            for m in data.get("modules", []):
                if m.get("api_supported", False): 
                    zoho_objects.append({
                        "name": m.get("api_name"),
                        "label": m.get("plural_label") or m.get("module_name") or m.get("api_name") 
                    })
            
            return sorted(zoho_objects, key=lambda x: x["label"])
    else:
        return [
            {"name": "account", "label": "Account (Mock)"},
            {"name": "contact", "label": "Contact (Mock)"}
        ]


# =========================================================
# ROUTE: FETCH SCHEMA FIELDS & SAMPLE PREVIEW DATA
# =========================================================
@router.get("/api/metadata/{crm_id}/fields/{object_name}")
async def get_crm_fields(
    crm_id: str,
    object_name: str,
    sf_token: str = Header(None, alias="sf-token"),
    sf_instance_url: str = Header(None, alias="sf-instance-url"),
    zd_token: str = Header(None, alias="zd-token"),
    zd_subdomain: str = Header(None, alias="zd-subdomain"),
    zoho_token: str = Header(None, alias="zoho-token"),            # Regional token header context
    zoho_api_domain: str = Header(None, alias="zoho-api-domain")  # Dynamic data center domain mapping
):
    crm_lower = crm_id.lower()
    
    if crm_lower == "salesforce":
        return await CrmMetadataService.fetch_salesforce_fields(sf_token, sf_instance_url, object_name)
        
    elif crm_lower == "zendesk":
        return await CrmMetadataService.fetch_zendesk_fields(zd_token, zd_subdomain, object_name)
    elif crm_lower == "zoho":
        # 1. Sanitize the Zoho Domain Protocol
        if zoho_api_domain and not zoho_api_domain.startswith(("http://", "https://")):
            zoho_api_domain = f"https://{zoho_api_domain}"
        base_url = zoho_api_domain.rstrip('/') if zoho_api_domain else "https://www.zohoapis.com"

        async with httpx.AsyncClient(verify=False, timeout=30.0) as client:
            headers = {"Authorization": f"Zoho-oauthtoken {zoho_token}"}

            # --- CALL 1: Fetch the Field Schema (Columns) ---
            fields_res = await client.get(
                f"{base_url}/crm/v6/settings/fields?module={object_name}", 
                headers=headers
            )
            
            if fields_res.status_code != 200:
                raise HTTPException(status_code=fields_res.status_code, detail="Failed to fetch field metadata schema from Zoho.")
                
            fields_data = fields_res.json().get("fields", [])
            
            # --- CALL 2: Fetch the Sample Data (Rows) ---
            records_res = await client.get(
                f"{base_url}/crm/v6/{object_name}?page=1&per_page=5", 
                headers=headers
            )
            
            sample_records = []
            if records_res.status_code == 200:
                raw_records = records_res.json().get("data", [])
                
                # Flatten complex Zoho data (like Owner or Lookup objects) into simple strings for the UI table
                for r in raw_records:
                    flat_rec = {}
                    for k, v in r.items():
                        if isinstance(v, dict) and "id" in v:
                            flat_rec[k] = v.get("name", v["id"]) 
                        else:
                            flat_rec[k] = v
                    sample_records.append(flat_rec)

            # Return both Schema AND Data to Angular
            return {
                "headers": [f["api_name"] for f in fields_data][:15], # Limit to first 15 headers for UI cleanliness
                "sampleRecords": sample_records,
                "fields": [
                    {
                        "name": f["api_name"],
                        "label": f["field_label"],
                        "type": f["data_type"],
                        "required": f.get("required", False)
                    } for f in fields_data
                ]
            }
       
    else:
        return {
            "headers": ["id", "name", "status"],
            "sampleRecords": [
                {"id": "MOCK-1", "name": "Global Test Enterprise", "status": "Active"},
                {"id": "MOCK-2", "name": "Acme Industrial Logistics", "status": "Inactive"}
            ],
            "fields": [
                {"name": "id", "label": "Record Identifier", "type": "string", "required": True},
                {"name": "name", "label": "Name Text", "type": "string", "required": True},
                {"name": "status", "label": "Status Flag", "type": "picklist", "required": False}
            ]
        }
    
@router.post("/api/metadata/preview-filter")
async def get_filtered_preview(request: Request):
    try:
        payload = await request.json()
        
        crm_id = payload.get("crmId", "").lower()
        obj_name = payload.get("objectName", "")
        query = payload.get("query", "").strip()
        headers_list = payload.get("headers", [])
        limit = int(payload.get("limit", 5))
        
        sf_token = payload.get("sfToken", "")
        sf_instance = payload.get("sfInstance", "")
        zd_token = payload.get("zdToken", "")
        zd_subdomain = payload.get("zdSubdomain", "")
        zoho_token = payload.get("zohoToken", "")
        zoho_api_domain = payload.get("zohoDomain", "")
        
        if not obj_name:
            raise HTTPException(status_code=400, detail="Object name is required.")

        # Using verify=False to bypass corporate proxy environment disruptions
        async with httpx.AsyncClient(verify=False, timeout=30.0) as client:
            
            # ==========================================
            # 1. SALESFORCE LIVE DYNAMIC QUERY
            # ==========================================
            if crm_id == "salesforce":
                if not sf_token or not sf_instance:
                    raise HTTPException(status_code=400, detail="Salesforce credentials missing.")
                
                # Trust the user's custom SOQL if provided, otherwise build it
                if query.lower().startswith("select "):
                    soql = query
                    if "limit " not in soql.lower():
                        soql += f" LIMIT {limit}"
                else:
                    fields_str = ", ".join(headers_list) if headers_list else "Id, Name"
                    where_clause = f" WHERE {query}" if query else ""
                    soql = f"SELECT {fields_str} FROM {obj_name}{where_clause} LIMIT {limit}"
                
                headers = {"Authorization": f"Bearer {sf_token}", "Content-Type": "application/json"}
                import urllib.parse
                safe_soql = urllib.parse.quote(soql)
                base_url = sf_instance.rstrip('/')
                url = f"{base_url}/services/data/v60.0/query?q={safe_soql}"
                
                res = await client.get(url, headers=headers)
                if res.status_code != 200:
                    err_msg = res.text
                    try:
                        err_data = res.json()
                        if isinstance(err_data, list) and len(err_data) > 0:
                            err_msg = err_data[0].get("message", res.text)
                    except: pass
                    raise HTTPException(status_code=400, detail=f"Salesforce rejected query: {err_msg.strip()}")
                    
                data = res.json()
                records = data.get("records", [])
                for r in records:
                    r.pop("attributes", None)
                    
                return {"records": records}

            # ==========================================
            # 2. ZENDESK LIVE DYNAMIC QUERY
            # ==========================================
            elif crm_id == "zendesk":
                if not zd_token or not zd_subdomain:
                    raise HTTPException(status_code=400, detail="Zendesk credentials missing.")
                
                safe_obj = obj_name.lower()
                headers = {"Authorization": f"Bearer {zd_token}", "Content-Type": "application/json"}
                
                # Check if it's a standard object
                standard_objects = ["tickets", "users", "organizations", "groups", "macros", "triggers", "views"]
                is_standard = safe_obj in standard_objects or f"{safe_obj}s" in standard_objects

                if is_standard:
                    # ==========================================
                    # STANDARD ZENDESK OBJECTS (Text Search)
                    # ==========================================
                    if safe_obj.endswith("s"): safe_obj = safe_obj[:-1]

                    # Strip out SQL junk if the frontend accidentally sent it to Zendesk
                    import re
                    clean_query = re.sub(r'(?i)^(select.*from\s+\w+\s+where\s+)', '', query).strip()

                    full_query = f"{clean_query} type:{safe_obj}" if clean_query else f"type:{safe_obj}"
                    import urllib.parse
                    safe_query = urllib.parse.quote(full_query)
                    
                    url = f"https://{zd_subdomain}.zendesk.com/api/v2/search.json?query={safe_query}&per_page={limit}"
                    
                    res = await client.get(url, headers=headers)
                    if res.status_code != 200:
                        raise HTTPException(status_code=400, detail=f"Zendesk Error: {res.text}")
                        
                    raw_records = res.json().get("results", [])
                    
                else:
                    # ==========================================
                    # NATIVE CUSTOM OBJECTS (JSON Filtered Search)
                    # ==========================================
                    if query.strip():
                        # Native Custom Object queries must be JSON filters, not text!
                        if not query.strip().startswith("{"):
                            raise HTTPException(
                                status_code=400, 
                                detail="Zendesk Custom Objects require a JSON Filter payload for queries (e.g., {\"filter\": {\"$and\": [...]}}). Leave the query blank to fetch recent records."
                            )
                        
                        import json
                        try:
                            json_payload = json.loads(query)
                        except json.JSONDecodeError:
                            raise HTTPException(status_code=400, detail="Invalid JSON payload in query.")
                            
                        # Use the POST search endpoint
                        url = f"https://{zd_subdomain}.zendesk.com/api/v2/custom_objects/{safe_obj}/records/search?page[size]={limit}"
                        res = await client.post(url, headers=headers, json=json_payload)
                    else:
                        # If query is empty, just list recent records via GET
                        url = f"https://{zd_subdomain}.zendesk.com/api/v2/custom_objects/{safe_obj}/records?page[size]={limit}"
                        res = await client.get(url, headers=headers)
                        
                    if res.status_code != 200:
                        raise HTTPException(status_code=400, detail=f"Zendesk Custom Object Error: {res.text}")
                        
                    raw_records = res.json().get("custom_object_records", [])

                # ==========================================
                # UNIFIED FLATTENING ENGINE
                # ==========================================
                flattened_records = []
                for rec in raw_records:
                    flat_rec = {}
                    for k, v in rec.items():
                        
                        # 1. Flatten Standard Object nested array
                        if k == "custom_fields" and isinstance(v, list):
                            for cf in v:
                                flat_rec[f"custom_field_{cf['id']}"] = cf.get("value")
                                
                        # 2. Flatten Native Custom Object nested dictionary
                        elif k == "custom_object_fields" and isinstance(v, dict):
                            for cf_key, cf_val in v.items():
                                flat_rec[cf_key] = cf_val
                                
                        # 3. Apply standard top-level fields (id, name, created_at)
                        elif not isinstance(v, (dict, list)): 
                            flat_rec[k] = v
                            
                    flattened_records.append(flat_rec)
                
                return {"records": flattened_records}
            
            # ==========================================
            # 3. ZOHO LIVE DYNAMIC QUERY
            # ==========================================
            elif crm_id == "zoho":
                if not zoho_token:
                    raise HTTPException(status_code=400, detail="Zoho authentication token context missing.")
                
                if zoho_api_domain and not zoho_api_domain.startswith(("http://", "https://")):
                    zoho_api_domain = f"https://{zoho_api_domain}"
                base_url = zoho_api_domain.rstrip('/') if zoho_api_domain else "https://www.zohoapis.com"
                
                headers = {"Authorization": f"Zoho-oauthtoken {zoho_token}"}
                
                if query:
                    coql_query = query.strip()
                    
                    # If user provides a full SELECT query, trust it but sanitize it
                    if coql_query.lower().startswith("select "):
                        
                        # 1. Replace '*' with actual headers (NO SPACES ALLOWED)
                        if "*" in coql_query:
                            clean_headers = [h for h in headers_list if not str(h).startswith("$")]
                            safe_fields = clean_headers[:40] if clean_headers else ["id"]
                            coql_query = coql_query.replace("*", ",".join(safe_fields), 1)
                            
                        # 2. ZOHO STRICT RULE: Erase all spaces after commas in the SELECT list
                        import re
                        match = re.match(r'(?i)select\s+(.*?)\s+from\s+', coql_query)
                        if match:
                            clean_select = match.group(1).replace(" ", "")
                            coql_query = coql_query.replace(match.group(1), clean_select, 1)

                        # 3. ZOHO STRICT RULE: A WHERE clause is 100% mandatory
                        # 3. ZOHO STRICT RULE: A WHERE clause is 100% mandatory
                        if " where " not in coql_query.lower():
                            if " limit " in coql_query.lower():
                                coql_query = re.sub(r'(?i)\s+limit\s+', ' where id is not null limit ', coql_query)
                            else:
                                coql_query += " where id is not null"

                        if " limit " not in coql_query.lower():
                            coql_query += f" limit {limit}"
                            
                    else:
                        # User only provided a WHERE clause
                        safe_fields = headers_list[:40] if headers_list else ["id"]
                        coql_query = f"select {','.join(safe_fields)} from {obj_name} where {coql_query} limit {limit}"

                    res = await client.post(
                        f"{base_url}/crm/v6/coql", 
                        headers=headers, 
                        json={"select_query": coql_query}
                    )
                else:
                    # Empty query fallback
                    safe_fields = headers_list[:40] if headers_list else ["id"]
                    fields_str = ",".join(safe_fields)
                    res = await client.get(
                        f"{base_url}/crm/v6/{obj_name}?page=1&per_page={limit}&fields={fields_str}", 
                        headers=headers
                    )

                # Accept both 200 (Success with data) and 204 (Success but 0 records found)
                if res.status_code not in [200, 204]:
                    print(f"\n{'='*50}\nZOHO RAW ERROR PAYLOAD:\n{res.text}\n{'='*50}\n")
                    
                    try:
                        err_data = res.json()
                        err_msg = str(err_data) 
                    except:
                        err_msg = res.text
                        
                    raise HTTPException(status_code=400, detail=f"Zoho rejected query: {err_msg}")
                    
                # Handle the 204 Empty Response safely
                if res.status_code == 204:
                    raw_records = []
                else:
                    raw_records = res.json().get("data") or []
                    
                sample_records = []
                
                for r in raw_records:
                    flat_rec = {}
                    for k, v in r.items():
                        if isinstance(v, dict) and "id" in v:
                            flat_rec[k] = v.get("name", v["id"]) 
                        else:
                            flat_rec[k] = v
                    
                    # Un-indented to properly append once per record!
                    sample_records.append(flat_rec)
                    
                return {"records": sample_records}                
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"System Crash: {e.__class__.__name__} - {str(e)}")