import json
import os
import tempfile
import shutil
import pandas as pd
from openpyxl import load_workbook
from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Request, Depends
from app.services.validator_service import process_validation_batch
from app.api.dependencies.auth import get_current_user
import glob
from datetime import datetime

router = APIRouter()

# ==========================================
# ROUTE 1: FAST HEADER EXTRACTION (CSV/EXCEL)
# ==========================================
@router.post("/api/python/extract-headers")
async def extract_headers(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename)[1].lower()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    temp_file_name = temp_file.name
    temp_file.close() 
    
    try:
        with open(temp_file_name, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        sheets = []
        headers_map = {}

        if ext == '.csv':
            df = pd.read_csv(temp_file_name, nrows=0)
            sheets = ["Sheet1"]
            headers_map["Sheet1"] = df.columns.tolist()
            
        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(temp_file_name, read_only=True, data_only=True)
            sheets = wb.sheetnames
            for sheet in sheets:
                ws = wb[sheet]
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
                headers_map[sheet] = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(first_row)]
            wb.close()
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format.")

        return {
            "sheets": sheets,
            "headersMap": headers_map
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        file.file.close()
        if os.path.exists(temp_file_name):
            try:
                os.remove(temp_file_name)
            except PermissionError:
                pass


# ==========================================
# ROUTE 2: MASSIVE DATA VALIDATION (CHUNKS)
# ==========================================
@router.post("/api/python/validate")
async def validate_batch(
    file: UploadFile = File(...),
    config: str = Form(...) 
):
    try:
        payload = json.loads(config)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid configuration format")

    mappings = payload.get("mappings", [])
    dedupe_key = payload.get("dedupeKey", "")
    sheet_name = payload.get("sheetName", "")
    sf_rules = payload.get("sfRules", {})
    date_format = payload.get("dateFormat", "")
    
    # ---  1: Extract Dynamic CRM ---
    target_crm = payload.get("targetCrmId", "salesforce")

    ext = os.path.splitext(file.filename)[1].lower()

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    temp_file_name = temp_file.name
    temp_file.close()

    try:
        with open(temp_file_name, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        total_count, total_valid, total_invalid_count, total_duplicates = 0, 0, 0, 0
        all_invalid_records, all_valid_records = [], []

        if ext == '.csv':
            chunk_iterator = pd.read_csv(temp_file_name, chunksize=10000)
            
            for chunk_df in chunk_iterator:
                chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                chunk_records = chunk_df.to_dict(orient="records")
                
                # --- 2: Pass target_rules and target_crm to the processor ---
                result = process_validation_batch(
                    records=chunk_records, mappings=mappings, dedupe_key=dedupe_key,  
                    target_rules=sf_rules, date_format=date_format, target_crm=target_crm
                )
                
                total_count += result["stats"]["total"]
                total_valid += result["stats"]["valid"]
                total_invalid_count += result["stats"]["invalid"]
                total_duplicates += result["stats"]["duplicates"]
                all_invalid_records.extend(result["invalidRecords"])
                all_valid_records.extend(result["validRecords"])

        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(temp_file_name, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            headers_raw = next(rows_iter, [])
            headers = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(headers_raw)]

            chunk_records = []
            
            for row in rows_iter:
                if not any(row): continue 
                
                chunk_records.append(dict(zip(headers, row)))
                
                if len(chunk_records) == 10000:
                    chunk_df = pd.DataFrame(chunk_records)
                    chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                    
                    # --- 3: Pass target_rules and target_crm to the processor ---
                    result = process_validation_batch(
                        records=chunk_df.to_dict(orient="records"), mappings=mappings, dedupe_key=dedupe_key,  
                        target_rules=sf_rules, date_format=date_format, target_crm=target_crm
                    )
                    
                    total_count += result["stats"]["total"]
                    total_valid += result["stats"]["valid"]
                    total_invalid_count += result["stats"]["invalid"]
                    total_duplicates += result["stats"]["duplicates"]
                    all_invalid_records.extend(result["invalidRecords"])
                    all_valid_records.extend(result["validRecords"])
                    
                    chunk_records = []
            
            if chunk_records:
                chunk_df = pd.DataFrame(chunk_records)
                chunk_df = chunk_df.astype(object).where(pd.notna(chunk_df), None)
                
                # --- 4: Pass target_rules and target_crm to the processor ---
                result = process_validation_batch(
                    records=chunk_df.to_dict(orient="records"), mappings=mappings, dedupe_key=dedupe_key,  
                    target_rules=sf_rules, date_format=date_format, target_crm=target_crm
                )
                
                total_count += result["stats"]["total"]
                total_valid += result["stats"]["valid"]
                total_invalid_count += result["stats"]["invalid"]
                total_duplicates += result["stats"]["duplicates"]
                all_invalid_records.extend(result["invalidRecords"])
                all_valid_records.extend(result["validRecords"])
                
            wb.close()

        return {
            "stats": {
                "total": total_count,
                "valid": total_valid,
                "invalid": total_invalid_count,
                "duplicates": total_duplicates
            },
            "invalidRecords": all_invalid_records,
            "validRecords": all_valid_records
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        file.file.close()
        if os.path.exists(temp_file_name):
            try:
                os.remove(temp_file_name)
            except PermissionError:
                pass


# ==========================================
# ROUTE 3: QUICK RE-VALIDATION (JSON)
# ==========================================
@router.post("/api/python/revalidate")
async def revalidate_batch_json(request: Request):
    payload = await request.json()
    records = payload.get("records", [])
    mappings = payload.get("mappings", [])
    dedupe_key = payload.get("dedupeKey", "")
    sf_rules = payload.get("sfRules", {})
    date_format = payload.get("dateFormat", "")
    
    # --- 5: Extract Dynamic CRM ---
    target_crm = payload.get("targetCrmId", "salesforce")

    # --- 6: Pass target_rules and target_crm to the processor ---
    result = process_validation_batch(
        records=records, 
        mappings=mappings, 
        dedupe_key=dedupe_key,  
        target_rules=sf_rules,
        date_format=date_format,
        target_crm=target_crm
    )
    
    return result

# ==========================================
# ROUTE 4: FETCH ACTIVE SESSIONS
# ==========================================
@router.get("/api/validation/sessions")
async def get_active_sessions(current_user = Depends(get_current_user)):
    """Scans the staging folder and returns a list of recoverable validation sessions.

    NOTE: staging .db files are organized on disk as {crm}/{object}/{session_id}.db
    with no user_id anywhere in that path, so this can only require *that you're
    logged in* -- it still can't tell one authenticated user's sessions apart from
    another's. Closing that fully needs a small schema addition (a
    validation_sessions table mapping session_id -> user_id, written at session
    creation time in migration_routes.py) so this endpoint -- and the audit
    download/revalidation routes -- can filter to sessions the caller actually
    owns instead of just requiring *a* valid login.
    """
    base_dir = os.path.join(os.getcwd(), "SureShift_staging_databases")
    sessions = []
    
    if not os.path.exists(base_dir):
        return {"sessions": sessions}
    
    for filepath in glob.glob(f"{base_dir}/*/*/*.db"):
        filename = os.path.basename(filepath)
        session_id = filename.replace('.db', '')
        parts = session_id.split('_')
        
        if len(parts) >= 4:
            crm = parts[0]
            obj = parts[1]
            date_str = f"{parts[2][:4]}-{parts[2][4:6]}-{parts[2][6:]} {parts[3][:2]}:{parts[3][2:4]}"
            
            stats = os.stat(filepath)
            size_mb = round(stats.st_size / (1024 * 1024), 2)
            
            sessions.append({
                "sessionId": session_id,
                "crm": crm.capitalize(),
                "object": obj.capitalize(),
                "date": date_str,
                "sizeMb": size_mb,
                "timestamp": stats.st_mtime
            })
            
    sessions.sort(key=lambda x: x["timestamp"], reverse=True)
    return {"sessions": sessions[:10]}